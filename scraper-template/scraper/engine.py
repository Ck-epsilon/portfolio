"""
Multi-Site Web Scraper — production-ready template.
Author: Ck.epsilon

Features:
- Playwright-based with stealth mode (anti-detection)
- Auto-pagination detection
- CSV & JSON export (CSV injection protected)
- Configurable via YAML
- Rate limiting with async sleep + retry with exponential backoff

Usage:
    python run.py --config sites/example.yaml --output data.csv
"""

import asyncio
import csv
import json
import logging
import random
import re
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urljoin

try:
    from playwright.async_api import async_playwright, Browser, BrowserContext, Page
except ImportError:
    print("pip install playwright && playwright install chromium")
    raise

logger = logging.getLogger(__name__)

# Characters that trigger formula execution in Excel/LibreOffice when at
# the start of a CSV cell. Prefix with single-quote to neutralize.
_CSV_DANGER_PREFIX = re.compile(r"^[=+\-@\t\r]")


@dataclass
class ScraperConfig:
    name: str
    start_urls: list[str]
    selectors: dict  # {field_name: css_selector}
    pagination_selector: str | None = None
    max_pages: int = 10
    delay_min: float = 1.0
    delay_max: float = 3.0
    max_retries: int = 3
    retry_backoff: float = 2.0
    user_agent: str = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    )
    output_format: str = "csv"  # csv | json | excel
    # Login config (optional — for authenticated scraping)
    login_url: str | None = None
    login_selector: str | None = None  # CSS selector for login form submit button
    login_credentials: dict[str, str] | None = None  # map field names → values


class StealthScraper:
    """Scraper with stealth features to avoid detection."""

    def __init__(self, config: ScraperConfig):
        self.config = config
        self.results: list[dict] = []
        self.browser: Browser | None = None

    async def __aenter__(self):
        self._playwright = await async_playwright().start()
        self.browser = await self._playwright.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
        return self

    async def __aexit__(self, *args):
        if self.browser:
            await self.browser.close()
        await self._playwright.stop()

    async def _new_context(self) -> BrowserContext:
        """Create a stealth-configured browser context."""
        return await self.browser.new_context(
            user_agent=self.config.user_agent,
            viewport={"width": 1920, "height": 1080},
            locale="en-US",
        )

    async def _stealth_page(self, context: BrowserContext) -> Page:
        """Create a page with anti-detection scripts."""
        page = await context.new_page()
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        """)
        return page

    async def _goto_with_retry(self, page: Page, url: str) -> None:
        """Navigate to URL with retry and exponential backoff."""
        last_exc = None
        for attempt in range(self.config.max_retries):
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                return
            except Exception as e:
                last_exc = e
                if attempt < self.config.max_retries - 1:
                    wait = self.config.retry_backoff ** attempt
                    logger.warning(
                        "Retry %d/%d for %s after %.1fs: %s",
                        attempt + 1, self.config.max_retries, url, wait, e,
                    )
                    await asyncio.sleep(wait)
        raise RuntimeError(
            f"Failed to load {url} after {self.config.max_retries} attempts"
        ) from last_exc

    async def _extract_items(self, page: Page) -> list[dict]:
        """Extract items from the current page using configured selectors."""
        primary_field = list(self.config.selectors.keys())[0]
        elements = await page.query_selector_all(
            self.config.selectors[primary_field]
        )

        if not elements:
            return []

        # Pre-fetch all other selector elements
        other_selectors = {}
        for field, sel in self.config.selectors.items():
            if field != primary_field:
                other_selectors[field] = await page.query_selector_all(sel)

        items = []
        for i, el in enumerate(elements):
            item = {primary_field: (await el.inner_text()).strip()}
            for field, els in other_selectors.items():
                if i < len(els):
                    item[field] = (await els[i].inner_text()).strip()
                else:
                    item[field] = ""
                    logger.debug(
                        "Field '%s' missing for item %d (expected %d elements, got %d)",
                        field, i, len(elements), len(els),
                    )
            items.append(item)
        return items

    async def scrape_page(self, url: str) -> list[dict]:
        """Scrape a single page."""
        context = await self._new_context()
        try:
            page = await self._stealth_page(context)
            await self._goto_with_retry(page, url)
            await page.wait_for_timeout(2000)
            return await self._extract_items(page)
        finally:
            await context.close()

    async def scrape_paginated(self, start_url: str) -> list[dict]:
        """Scrape with automatic pagination — single context, single page reused."""
        all_items = []
        context = await self._new_context()
        try:
            page = await self._stealth_page(context)
            current_url = start_url

            for _ in range(self.config.max_pages):
                await self._goto_with_retry(page, current_url)
                await page.wait_for_timeout(1500)

                items = await self._extract_items(page)
                all_items.extend(items)

                next_btn = await page.query_selector(
                    self.config.pagination_selector
                )
                if not next_btn:
                    break

                href = await next_btn.get_attribute("href")
                if not href:
                    break

                current_url = urljoin(current_url, href)
                delay = random.uniform(
                    self.config.delay_min, self.config.delay_max
                )
                await asyncio.sleep(delay)
        finally:
            await context.close()

        return all_items

    async def login(self, context: BrowserContext) -> Page:
        """Handle login flow: navigate to login page, fill credentials, submit.
        Returns an authenticated Page ready for scraping."""
        if not self.config.login_url or not self.config.login_credentials:
            raise ValueError("login_url and login_credentials required for login")
        page = await self._stealth_page(context)
        await self._goto_with_retry(page, self.config.login_url)
        await page.wait_for_timeout(1500)
        for field_name, value in self.config.login_credentials.items():
            await page.fill(f'input[name="{field_name}"]', value)
        if self.config.login_selector:
            await page.click(self.config.login_selector)
        else:
            await page.click('button[type="submit"]')
        await page.wait_for_timeout(3000)
        logger.info("Login flow completed")
        return page

    async def run(self) -> list[dict]:
        """Execute full scrape across all configured URLs."""
        if self.config.login_url:
            print(f"[{self.config.name}] Logging in via {self.config.login_url}...")
        print(f"[{self.config.name}] Starting scrape...")

        for url in self.config.start_urls:
            print(f"  → {url}")

            try:
                if self.config.pagination_selector:
                    items = await self.scrape_paginated(url)
                else:
                    items = await self.scrape_page(url)
            except Exception as e:
                logger.error("Failed to scrape %s: %s", url, e)
                print(f"    ⚠ Skipped: {e}")
                continue

            self.results.extend(items)
            print(f"    {len(items)} items extracted")

            delay = random.uniform(self.config.delay_min, self.config.delay_max)
            await asyncio.sleep(delay)

        print(f"[{self.config.name}] Done. Total: {len(self.results)} items")
        return self.results

    def _sanitize_csv_cell(self, value: str) -> str:
        """Prefix dangerous characters to prevent CSV formula injection."""
        if _CSV_DANGER_PREFIX.match(value):
            return "'" + value
        return value

    def save(self, output_path: str):
        """Save results to CSV (injection-safe), JSON, or Excel."""
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        if self.config.output_format == "excel":
            self.save_excel(str(output))
            return

        if self.config.output_format == "json":
            with open(output, "w", encoding="utf-8") as f:
                json.dump(self.results, f, ensure_ascii=False, indent=2)
        else:
            if not self.results:
                print("Warning: No results to save.")
                return
            fieldnames = list(self.results[0].keys())
            with open(output, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                sanitized = [
                    {k: self._sanitize_csv_cell(str(v)) for k, v in row.items()}
                    for row in self.results
                ]
                writer.writerows(sanitized)

        print(f"Saved {len(self.results)} items → {output.absolute()}")

    def save_db(self, db_path: str = "scraper.db", table_name: str = "results"):
        """Save results to SQLite database. Creates table if not exists."""
        import sqlite3
        if not self.results:
            print("Warning: No results to save to DB.")
            return
        conn = sqlite3.connect(db_path)
        try:
            fieldnames = list(self.results[0].keys())
            cols = ", ".join(f'"{f}" TEXT' for f in fieldnames)
            conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({cols})')
            placeholders = ", ".join("?" for _ in fieldnames)
            rows = [[str(row.get(f, "")) for f in fieldnames] for row in self.results]
            conn.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders})', rows)
            conn.commit()
            print(f"Saved {len(self.results)} items to SQLite → {db_path}::{table_name}")
        finally:
            conn.close()

    def save_excel(self, output_path: str):
        """Save results to a formatted .xlsx file using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise

        if not self.results:
            print("Warning: No results to save.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.config.name[:31]  # Excel sheet name max 31 chars

        fieldnames = list(self.results[0].keys())

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center")

        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows with CSV injection protection
        for row_idx, item in enumerate(self.results, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = str(item.get(field, ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=self._sanitize_csv_cell(value))
                # Auto-width: max(current width, len(value) * 1.2)
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(
                    ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width or 0,
                    len(value) * 1.2 + 2,
                )

        wb.save(output_path)
        print(f"Saved {len(self.results)} items to Excel → {output_path}")

    def clean_data(
        self,
        dedup_keys: list[str] | None = None,
        normalize_fields: list[str] | None = None,
        required_fields: list[str] | None = None,
    ) -> list[dict]:
        """Clean scraped data: deduplicate, normalize whitespace, validate required fields.
        Returns the cleaned subset. Also stores back into self.results."""
        cleaned = self.results[:]

        # Deduplicate by specified keys
        if dedup_keys:
            seen = set()
            deduped = []
            for row in cleaned:
                key = tuple(row.get(k, "") for k in dedup_keys)
                if key not in seen:
                    seen.add(key)
                    deduped.append(row)
            cleaned = deduped

        # Normalize whitespace on specified fields
        if normalize_fields:
            for row in cleaned:
                for field in normalize_fields:
                    if field in row and isinstance(row[field], str):
                        row[field] = " ".join(row[field].split())

        # Validate required fields (drop rows with empty required fields)
        if required_fields:
            before = len(cleaned)
            cleaned = [row for row in cleaned if all(row.get(f, "").strip() for f in required_fields)]
            logger.info("Validation dropped %d rows (missing required fields)", before - len(cleaned))

        self.results = cleaned
        return cleaned
